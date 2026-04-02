#!/usr/bin/env python3
"""
TVT NVR Camera Scanner

Reads a JSON file of NVR devices, connects to each TVT NVR via the native
binary protocol (port 9008), and retrieves all programmed camera information.
The SDK backend talks to a Fastify API running inside Docker (or natively on
Linux) which wraps the native TVT SDK.

Usage:
    python main.py <nvr_devices.json> [--config config.json] [--output results.csv]
    python main.py <nvr_devices.json> --site "Site A"
    python main.py <nvr_devices.json> --output results.json
    python main.py <nvr_devices.json> --backend sdk --api-url http://localhost:3000
"""

import dotenv
import argparse
import csv
import json
import os
import subprocess
import sys
import urllib.request
import urllib.error
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

dotenv.load_dotenv()
SCRIPT_DIR = Path(__file__).parent.resolve()

from tvt_protocol import scan_nvr as python_scan_nvr

# Default API URL for the TVT SDK Fastify server
DEFAULT_API_URL = os.getenv("TVT_API_URL", "http://localhost:3000")

# Path to the standalone scan_nvr.mjs script (for sdk-local backend)
SCAN_SCRIPT = SCRIPT_DIR / "scan_nvr.mjs"


def load_config(config_path: str | None) -> dict:
    defaults = {
        "username": os.getenv("TVT_USERNAME", "admin"),
        "password": os.getenv("TVT_PASSWORD", ""),
        "port": int(os.getenv("TVT_PORT", 6036)),
        "timeout": int(os.getenv("TVT_TIMEOUT", 10)),
        "max_channels": int(os.getenv("TVT_MAX_CHANNELS", 64)),
        "concurrency": int(os.getenv("TVT_CONCURRENCY", 4)),
    }
    if config_path and os.path.exists(config_path):
        with open(config_path) as f:
            user_config = json.load(f)
        defaults.update(user_config)
    return defaults


def load_devices(json_path: str) -> list[dict]:
    with open(json_path) as f:
        devices = json.load(f)

    if not isinstance(devices, list):
        print(f"Error: {json_path} must contain a JSON array", file=sys.stderr)
        sys.exit(1)

    return devices


def filter_tvt_devices(devices: list[dict]) -> list[dict]:
    """Filter to only TVT manufacturer devices (or devices with TVT MAC prefix 58:5B:69)."""
    tvt_devices = []
    for d in devices:
        manufacturer = (d.get("manufacturer") or "").strip().upper()
        mac = (d.get("mac") or "").strip().upper()
        # Include if manufacturer is TVT/UNIQUE or MAC starts with TVT OUI
        if manufacturer in ("TVT", "UNIQUE") or mac.startswith("58:5B:69"):
            tvt_devices.append(d)
    return tvt_devices


def scan_single_nvr(device: dict, config: dict, backend: str = "protocol") -> dict:
    """Scan a single NVR using the chosen backend.

    backend = "protocol"   -> pure Python binary protocol on port 9008
    backend = "sdk"        -> HTTP API (Fastify server in Docker)
    backend = "sdk-local"  -> subprocess calling scan_nvr.mjs (needs Node.js + SDK on Linux)
    backend = "both"       -> try protocol first, fall back to sdk
    """
    if backend == "protocol":
        return _protocol_scan(device, config)
    elif backend == "sdk":
        return _sdk_scan(device, config)
    elif backend == "sdk-local":
        return _sdk_scan_local(device, config)
    elif backend == "both":
        result = _protocol_scan(device, config)
        if not result["success"]:
            sdk_result = _sdk_scan(device, config)
            if sdk_result["success"]:
                return sdk_result
            result["error"] = f"protocol: {result['error']} | sdk: {sdk_result['error']}"
        return result
    else:
        raise ValueError(f"Unknown backend: {backend}")


def _protocol_scan(device: dict, config: dict) -> dict:
    """Scan a single NVR using the pure Python TVT binary protocol."""
    ip = device["ip"]
    port = device.get("port", config["port"])
    username = config["username"]
    password = config["password"]
    timeout = config["timeout"]

    scan_result = python_scan_nvr(ip, port=port, username=username,
                                   password=password, timeout=timeout)

    return {
        "site": device.get("site", ""),
        "hostname": device.get("hostname", ""),
        "nvr_ip": ip,
        "nvr_mac": device.get("mac", ""),
        "nvr_port": port,
        "success": scan_result.get("success", False),
        "device_name": scan_result.get("device_name", ""),
        "device_model": scan_result.get("device_model", ""),
        "serial_number": scan_result.get("serial_number", ""),
        "firmware": scan_result.get("firmware", ""),
        "total_channels": scan_result.get("total_channels", 0),
        "cameras": scan_result.get("cameras", []),
        "error": scan_result.get("error"),
        "backend": "protocol",
        "device_info": scan_result.get("device_info", {}),
    }


def _sdk_scan_local(device: dict, config: dict) -> dict:
    """Scan a single NVR via subprocess calling scan_nvr.mjs (needs Node.js + SDK on Linux)."""
    ip = device["ip"]
    port = device.get("port", config["port"])
    username = config["username"]
    password = config["password"]

    result = {
        "site": device.get("site", ""),
        "hostname": device.get("hostname", ""),
        "nvr_ip": ip,
        "nvr_mac": device.get("mac", ""),
        "nvr_port": port,
        "success": False,
        "device_name": "",
        "device_model": "",
        "serial_number": "",
        "firmware": "",
        "total_channels": 0,
        "cameras": [],
        "error": None,
        "backend": "sdk-local",
    }

    if not SCAN_SCRIPT.exists():
        result["error"] = f"scan_nvr.mjs not found at {SCAN_SCRIPT}"
        return result

    try:
        proc = subprocess.run(
            ["node", str(SCAN_SCRIPT), ip, str(port), username, password],
            capture_output=True, text=True,
            timeout=config["timeout"] + 30,
        )

        stdout = proc.stdout
        # Extract JSON between sentinel markers (SDK may pollute stdout)
        start_marker = "___JSON_START___"
        end_marker = "___JSON_END___"
        start_idx = stdout.find(start_marker)
        end_idx = stdout.find(end_marker)

        if start_idx == -1 or end_idx == -1:
            result["error"] = f"No JSON markers in output. stderr: {proc.stderr[:500]}"
            return result

        json_str = stdout[start_idx + len(start_marker):end_idx].strip()
        scan_data = json.loads(json_str)

        result.update({
            "success": scan_data.get("success", False),
            "device_name": scan_data.get("device_name", ""),
            "device_model": scan_data.get("device_model", ""),
            "serial_number": scan_data.get("serial_number", ""),
            "firmware": scan_data.get("firmware", ""),
            "total_channels": scan_data.get("total_channels", 0),
            "cameras": scan_data.get("cameras", []),
            "error": scan_data.get("error"),
        })

    except subprocess.TimeoutExpired:
        result["error"] = f"Subprocess timeout after {config['timeout'] + 30}s"
    except json.JSONDecodeError as e:
        result["error"] = f"Invalid JSON from scan_nvr.mjs: {e}"
    except FileNotFoundError:
        result["error"] = "node not found in PATH"

    return result


def _sdk_scan(device: dict, config: dict) -> dict:
    """Scan a single NVR by calling the TVT SDK Fastify API."""
    ip = device["ip"]
    port = device.get("port", config["port"])
    username = config["username"]
    password = config["password"]
    api_url = config.get("api_url", DEFAULT_API_URL)

    result = {
        "site": device.get("site", ""),
        "hostname": device.get("hostname", ""),
        "nvr_ip": ip,
        "nvr_mac": device.get("mac", ""),
        "nvr_port": port,
        "success": False,
        "device_name": "",
        "device_model": "",
        "serial_number": "",
        "firmware": "",
        "total_channels": 0,
        "cameras": [],
        "error": None,
    }

    try:
        payload = json.dumps({
            "ip": ip,
            "port": port,
            "username": username,
            "password": password,
            "maxCameras": config.get("max_channels", 64),
        }).encode()

        req = urllib.request.Request(
            f"{api_url}/scan",
            data=payload,
            headers={"Content-Type": "application/json"},
            method="POST",
        )
        timeout = config["timeout"] + 15

        with urllib.request.urlopen(req, timeout=timeout) as resp:
            scan_data = json.loads(resp.read().decode())

        result.update({
            "success": scan_data.get("success", False),
            "device_name": scan_data.get("device_name", ""),
            "device_model": scan_data.get("device_model", ""),
            "serial_number": scan_data.get("serial_number", ""),
            "firmware": scan_data.get("firmware", ""),
            "total_channels": scan_data.get("total_channels", 0),
            "cameras": scan_data.get("cameras", []),
            "error": scan_data.get("error"),
            "backend": "sdk",
        })

    except urllib.error.URLError as e:
        result["error"] = f"API connection error: {e.reason}"
    except TimeoutError:
        result["error"] = f"API timeout after {config['timeout'] + 15}s"
    except json.JSONDecodeError as e:
        result["error"] = f"Invalid JSON from API: {e}"

    return result


def print_nvr_report(nvr_result: dict) -> None:
    """Print a formatted report for a single NVR."""
    site = nvr_result["site"]
    hostname = nvr_result["hostname"]
    ip = nvr_result["nvr_ip"]
    success = nvr_result["success"]

    header = f"\n{'='*80}"
    print(header)
    print(f"  Site: {site}")
    print(f"  NVR:  {hostname} ({ip})")
    print(f"{'='*80}")

    if not success:
        print(f"  ERROR: {nvr_result.get('error', 'Unknown error')}")
        return

    dev_name = nvr_result.get("device_name", "")
    dev_model = nvr_result.get("device_model", "")
    firmware = nvr_result.get("firmware", "")
    sn = nvr_result.get("serial_number", "")
    total = nvr_result.get("total_channels", 0)

    if dev_name or dev_model:
        print(f"  Device: {dev_name} | Model: {dev_model}")
    if firmware:
        print(f"  Firmware: {firmware}")
    if sn:
        print(f"  S/N: {sn}")
    print(f"  Total Channels: {total}")
    print(f"  {'-'*76}")

    cameras = nvr_result.get("cameras", [])
    cameras = [c for c in cameras if c.get("address", "").strip()]
    if cameras:
        print(f"  {'Ch':<4} {'Camera Name':<28} {'Address':<18} {'Port':<6} {'Status':<8} {'Model'}")
        print(f"  {'--':<4} {'-'*27:<28} {'-'*17:<18} {'----':<6} {'------':<8} {'-----'}")
        for cam in cameras:
            ch = cam.get("channel", "?")
            name = cam.get("name", "")[:27]
            addr = cam.get("address", "")
            port = cam.get("port", "")
            status = cam.get("status", "?")
            model = cam.get("model", "")
            print(f"  {ch:<4} {name:<28} {addr:<18} {port:<6} {status:<8} {model}")
    else:
        print("  No camera details available")


def save_csv(all_results: list[dict], output_path: str) -> None:
    """Save all camera data to CSV."""
    rows = []
    for nvr in all_results:
        cameras = nvr.get("cameras", [])
        if cameras:
            for cam in cameras:
                if not cam.get("address", "").strip():
                    continue
                rows.append({
                    "site": nvr["site"],
                    "nvr_hostname": nvr["hostname"],
                    "nvr_ip": nvr["nvr_ip"],
                    "nvr_mac": nvr["nvr_mac"],
                    "nvr_device_name": nvr.get("device_name", ""),
                    "nvr_model": nvr.get("device_model", ""),
                    "nvr_firmware": nvr.get("firmware", ""),
                    "nvr_serial": nvr.get("serial_number", ""),
                    "channel": cam.get("channel", ""),
                    "camera_name": cam.get("name", ""),
                    "camera_address": cam.get("address", ""),
                    "camera_port": cam.get("port", ""),
                    "camera_status": cam.get("status", ""),
                    "camera_protocol": cam.get("protocol", ""),
                    "camera_model": cam.get("model", ""),
                })
        else:
            # Still record the NVR even if no cameras retrieved
            rows.append({
                "site": nvr["site"],
                "nvr_hostname": nvr["hostname"],
                "nvr_ip": nvr["nvr_ip"],
                "nvr_mac": nvr["nvr_mac"],
                "nvr_device_name": nvr.get("device_name", ""),
                "nvr_model": nvr.get("device_model", ""),
                "nvr_firmware": nvr.get("firmware", ""),
                "nvr_serial": nvr.get("serial_number", ""),
                "channel": "",
                "camera_name": "",
                "camera_address": "",
                "camera_port": "",
                "camera_status": "",
                "camera_protocol": "",
                "camera_model": "",
                "error": nvr.get("error", ""),
            })

    if not rows:
        print("No data to save.", file=sys.stderr)
        return

    fieldnames = list(rows[0].keys())
    with open(output_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    print(f"\nCSV saved to: {output_path}")


def save_json(all_results: list[dict], output_path: str) -> None:
    """Save all results to JSON."""
    with open(output_path, "w") as f:
        json.dump(all_results, f, indent=2)
    print(f"\nJSON saved to: {output_path}")


def _xlsx_upsert_sheet(wb, sheet_name):
    """Remove a sheet if it exists and create a fresh one."""
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    return wb.create_sheet(sheet_name)


def _xlsx_write_header(ws, columns, header_font, header_fill, header_align, thin_border):
    """Write a styled header row."""
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border


def _xlsx_autofit(ws, columns, last_row):
    """Auto-fit column widths based on content."""
    from openpyxl.utils import get_column_letter
    for col_idx in range(1, len(columns) + 1):
        max_len = len(columns[col_idx - 1])
        for row in range(2, last_row + 1):
            val = ws.cell(row=row, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 50)


def save_xlsx_per_site(all_results: list[dict], output_dir: str) -> None:
    """Save scan results as one XLSX file per site with NVR Config and NVR Info tabs.

    Each site gets its own .xlsx file in output_dir. The file is created if it
    does not exist; existing tabs are replaced (upserted) on re-run.

    Tabs created:
      - NVR Config: one row per camera (only cameras with an address)
      - NVR Info:   one row per NVR with device-level details
    """
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    # Group results by site
    sites: dict[str, list[dict]] = defaultdict(list)
    for result in all_results:
        site_name = result.get("site", "Unknown Site")
        sites[site_name].append(result)

    # Style definitions
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    cam_columns = ["Camera Name", "Address", "Port", "Status", "Protocol", "Model"]
    nvr_columns = [
        "Hostname", "IP", "MAC", "Port", "Device Name", "Model",
        "Serial Number", "Firmware", "Channels", "Status",
    ]

    created_files = []

    for site_name, nvr_results in sorted(sites.items()):
        # Sanitize filename
        safe_name = "".join(c if c.isalnum() or c in " -_#()" else "_" for c in site_name).strip()
        xlsx_path = output_path / f"{safe_name}.xlsx"

        # Load existing workbook or create new one
        if xlsx_path.exists():
            wb = load_workbook(xlsx_path)
        else:
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

        # ── NVR Info tab ─────────────────────────────────────────────
        ws_info = _xlsx_upsert_sheet(wb, "NVR Info")
        _xlsx_write_header(ws_info, nvr_columns, header_font, header_fill, header_align, thin_border)

        info_row = 2
        for nvr in sorted(nvr_results, key=lambda r: r.get("hostname", "")):
            status = "OK" if nvr["success"] else nvr.get("error", "Failed")
            values = [
                nvr.get("hostname", ""),
                nvr.get("nvr_ip", ""),
                nvr.get("nvr_mac", ""),
                nvr.get("nvr_port", ""),
                nvr.get("device_name", ""),
                nvr.get("device_model", ""),
                nvr.get("serial_number", ""),
                nvr.get("firmware", ""),
                nvr.get("total_channels", 0),
                status,
            ]
            for col_idx, value in enumerate(values, 1):
                cell = ws_info.cell(row=info_row, column=col_idx, value=value)
                cell.border = thin_border
            info_row += 1

        _xlsx_autofit(ws_info, nvr_columns, info_row - 1)
        if info_row > 2:
            ws_info.auto_filter.ref = f"A1:{get_column_letter(len(nvr_columns))}{info_row - 1}"

        # ── NVR Config (cameras) tab ─────────────────────────────────
        ws_cam = _xlsx_upsert_sheet(wb, "NVR Config")
        _xlsx_write_header(ws_cam, cam_columns, header_font, header_fill, header_align, thin_border)

        cam_row = 2
        for nvr in nvr_results:
            for cam in nvr.get("cameras", []):
                cam_addr = cam.get("address", "").strip()
                if not cam_addr:
                    continue
                values = [
                    cam.get("name", ""),
                    cam_addr,
                    cam.get("port", ""),
                    cam.get("status", ""),
                    cam.get("protocol", ""),
                    cam.get("model", ""),
                ]
                for col_idx, value in enumerate(values, 1):
                    cell = ws_cam.cell(row=cam_row, column=col_idx, value=value)
                    cell.border = thin_border
                cam_row += 1

        _xlsx_autofit(ws_cam, cam_columns, cam_row - 1)
        if cam_row > 2:
            ws_cam.auto_filter.ref = f"A1:{get_column_letter(len(cam_columns))}{cam_row - 1}"

        # ── Save ─────────────────────────────────────────────────────
        wb.save(xlsx_path)
        cam_count = cam_row - 2
        nvr_count = info_row - 2
        created_files.append((site_name, str(xlsx_path), nvr_count, cam_count))

    print(f"\nXLSX files saved to: {output_dir}")
    for site_name, fpath, nvrs, cams in created_files:
        print(f"  {site_name}: {nvrs} NVR(s), {cams} cameras -> {fpath}")


def save_failed_devices(all_results: list[dict], devices: list[dict], output_path: str) -> None:
    """Write a JSON file containing devices that failed to scan.

    Matches failed results back to the original device entries so the output
    file can be fed directly back into the scanner for a retry.
    """
    failed_ips = {r["nvr_ip"] for r in all_results if not r["success"]}
    # Also include devices that were in the input but produced no result at all
    scanned_ips = {r["nvr_ip"] for r in all_results}
    missing_ips = {d["ip"] for d in devices} - scanned_ips

    all_failed_ips = failed_ips | missing_ips
    if not all_failed_ips:
        return

    failed_devices = [d for d in devices if d["ip"] in all_failed_ips]
    with open(output_path, "w") as f:
        json.dump(failed_devices, f, indent=2)

    print(f"\n{len(failed_devices)} failed/missing device(s) saved to: {output_path}")


def main():
    parser = argparse.ArgumentParser(
        description="TVT NVR Camera Scanner - Retrieve camera lists from TVT NVRs"
    )
    parser.add_argument(
        "input",
        help="Path to JSON file with NVR device list",
    )
    parser.add_argument(
        "--config",
        default=str(SCRIPT_DIR / "config.json"),
        help="Path to config.json (default: config.json in script dir)",
    )
    parser.add_argument(
        "--output", "-o",
        help="Output file path (.csv or .json)",
    )
    parser.add_argument(
        "--site", "-s",
        help="Filter to a specific site name (partial match)",
    )
    parser.add_argument(
        "--concurrency", "-c",
        type=int,
        help="Max concurrent NVR scans (default: from config)",
    )
    parser.add_argument(
        "--username", "-u",
        help="NVR username (overrides config)",
    )
    parser.add_argument(
        "--password", "-p",
        help="NVR password (overrides config)",
    )
    parser.add_argument(
        "--backend", "-b",
        choices=["protocol", "sdk", "sdk-local", "both"],
        default="protocol",
        help="Scan backend: protocol (pure Python, default), sdk (HTTP API), sdk-local (subprocess), both (protocol then sdk)",
    )
    parser.add_argument(
        "--xlsx",
        metavar="DIR",
        help="Output one .xlsx file per site into DIR (e.g. --xlsx files/)",
    )
    parser.add_argument(
        "--failed",
        metavar="FILE",
        help="Write devices that failed to scan to a JSON file for retry (e.g. --failed failed_devices.json)",
    )
    parser.add_argument(
        "--api-url",
        default=DEFAULT_API_URL,
        help=f"TVT SDK API URL for sdk backend (default: {DEFAULT_API_URL})",
    )
    args = parser.parse_args()

    # Load config
    config = load_config(args.config)
    config["api_url"] = args.api_url
    if args.concurrency:
        config["concurrency"] = args.concurrency
    if args.username:
        config["username"] = args.username
    if args.password:
        config["password"] = args.password

    # Load and filter devices
    all_devices = load_devices(args.input)
    tvt_devices = filter_tvt_devices(all_devices)

    if args.site:
        site_filter = args.site.lower()
        tvt_devices = [d for d in tvt_devices if site_filter in d.get("site", "").lower()]

    # De-duplicate by IP (same NVR can appear in multiple sites)
    seen_ips = set()
    unique_devices = []
    for d in tvt_devices:
        ip = d["ip"]
        # Skip link-local addresses
        if ip.startswith("169.254."):
            continue
        if ip not in seen_ips:
            seen_ips.add(ip)
            unique_devices.append(d)

    if not unique_devices:
        print("No TVT devices found matching the criteria.", file=sys.stderr)
        sys.exit(1)

    print(f"\nTVT NVR Camera Scanner")
    print(f"Found {len(unique_devices)} unique TVT NVR(s) to scan")
    print(f"Backend: {args.backend}")
    print(f"Credentials: {config['username']} / {'*' * len(config['password'])}")
    print(f"Concurrency: {config['concurrency']}")

    # Scan all NVRs
    all_results = []
    concurrency = config["concurrency"]
    backend = args.backend

    with ThreadPoolExecutor(max_workers=concurrency) as executor:
        future_to_device = {
            executor.submit(scan_single_nvr, device, config, backend): device
            for device in unique_devices
        }

        for i, future in enumerate(as_completed(future_to_device), 1):
            device = future_to_device[future]
            try:
                result = future.result()
                all_results.append(result)
                status = "OK" if result["success"] else "FAIL"
                cam_count = sum(1 for c in result.get("cameras", []) if c.get("address", "").strip())
                print(f"  [{i}/{len(unique_devices)}] {device['site']} / {device.get('hostname', device['ip'])} - {status} ({cam_count} cameras)")
            except Exception as e:
                print(f"  [{i}/{len(unique_devices)}] {device['site']} / {device['ip']} - ERROR: {e}", file=sys.stderr)

    # Sort results by site name
    all_results.sort(key=lambda r: r.get("site", ""))

    # Print reports
    for result in all_results:
        print_nvr_report(result)

    # Summary
    total_nvrs = len(all_results)
    successful = sum(1 for r in all_results if r["success"])
    total_cameras = sum(
        1 for r in all_results for c in r.get("cameras", [])
        if c.get("address", "").strip()
    )
    print(f"\n{'='*80}")
    print(f"  SUMMARY: {successful}/{total_nvrs} NVRs scanned successfully, {total_cameras} cameras with addresses found")
    print(f"{'='*80}")

    # Save output
    if args.output:
        output_path = args.output
        if output_path.endswith(".csv"):
            save_csv(all_results, output_path)
        else:
            save_json(all_results, output_path)

    if args.xlsx:
        save_xlsx_per_site(all_results, args.xlsx)

    # Always save failed devices if any failed
    failed_path = args.failed or str(SCRIPT_DIR / "files" / "failed_devices.json")
    save_failed_devices(all_results, unique_devices, failed_path)


if __name__ == "__main__":
    main()
