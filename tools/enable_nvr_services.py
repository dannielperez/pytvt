#!/usr/bin/env python3
"""Batch-enable RTSP + API Server on all TVT NVRs from ruijie inventory.

Reads data/results/ruijie/nvr_devices.json, filters out:
  - 192.168.110.0/24 subnet
  - Hikvision / Dahua / Fiberhome manufacturers
  - Non-NVR hostnames (e.g. "Transfer Server", "Speaker")

For each remaining NVR:
  1. Probe HTTP port (try 80, then 8080, then 443)
  2. Attempt NVR web API login
  3. Enable RTSP + API Server if not already enabled
  4. Record result

Saves report to data/results/tvt/nvr_service_report.json (and .xlsx if openpyxl available).
"""

import json
import os
import socket
import sys
import time
from datetime import datetime

# Add parent for imports
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from pytvt.xml_api import NvrClient
from pytvt.models import NvrApiError

INVENTORY_PATH = os.path.join(os.path.dirname(__file__), "..", "..", "data", "results", "ruijie", "nvr_devices.json")
REPORT_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "data", "results", "tvt")
NVR_PASSWORD = os.getenv("NVR_PASSWORD", "")

# Subnets / manufacturers to skip
SKIP_SUBNETS = ["192.168.110."]
SKIP_MANUFACTURERS = {"hikvision", "dahua", "fiberhome"}
SKIP_HOSTNAMES = {"transfer server", "speaker"}
HTTP_PORTS_TO_TRY = [80, 8080, 443]


def is_port_open(ip: str, port: int, timeout: float = 3.0) -> bool:
    """Quick TCP connect check."""
    try:
        with socket.create_connection((ip, port), timeout=timeout):
            return True
    except (OSError, socket.timeout):
        return False


def find_http_port(ip: str) -> int | None:
    """Find the NVR's HTTP port by trying common ports."""
    for port in HTTP_PORTS_TO_TRY:
        if is_port_open(ip, port):
            return port
    return None


def should_skip(device: dict) -> str | None:
    """Return skip reason or None if device should be processed."""
    ip = device.get("ip", "")
    mfr = device.get("manufacturer", "").lower().strip()
    hostname = device.get("hostname", "").lower().strip()

    for subnet in SKIP_SUBNETS:
        if ip.startswith(subnet):
            return f"subnet {subnet}0/24"

    if mfr in SKIP_MANUFACTURERS:
        return f"manufacturer: {mfr}"

    for skip_name in SKIP_HOSTNAMES:
        if skip_name in hostname:
            return f"hostname: {hostname}"

    return None


def process_nvr(ip: str, port: int, username: str, password: str, timeout: int = 10) -> dict:
    """Try to login and enable services on a single NVR."""
    result = {
        "http_port": port,
        "reachable": True,
        "login_ok": False,
        "rtsp_was_enabled": None,
        "api_was_enabled": None,
        "rtsp_enabled": False,
        "api_enabled": False,
        "rtsp_port": None,
        "channels": None,
        "error": None,
    }

    try:
        with NvrClient(ip, username, password, port=port, timeout=timeout) as nvr:
            nvr.login()
            result["login_ok"] = True

            # Query current state
            rtsp = nvr.query_rtsp_server()
            api = nvr.query_api_server()
            result["rtsp_was_enabled"] = rtsp.enabled
            result["api_was_enabled"] = api.enabled
            result["rtsp_port"] = rtsp.port

            # Enable if needed
            changed = nvr.ensure_services_enabled()
            result["rtsp_enabled"] = True
            result["api_enabled"] = True

            # Get channel count
            try:
                channels = nvr.query_channels()
                result["channels"] = len(channels)
            except Exception:
                pass

            if changed:
                result["action"] = "enabled: " + ", ".join(changed.keys())
            else:
                result["action"] = "already enabled"

    except NvrApiError as e:
        result["error"] = str(e)
    except Exception as e:
        result["error"] = f"{type(e).__name__}: {e}"

    return result


def save_xlsx(report: list[dict], path: str) -> bool:
    """Save report to Excel if openpyxl is available."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "NVR Service Report"

    # Header
    headers = ["Site", "IP", "Hostname", "Manufacturer", "HTTP Port",
               "Reachable", "Login OK", "RTSP Was On", "API Was On",
               "RTSP Now", "API Now", "RTSP Port", "Channels", "Action", "Error", "Status"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Status colors
    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    gray = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    for row_idx, r in enumerate(report, 2):
        status = r.get("status", "")
        vals = [
            r.get("site", ""), r.get("ip", ""), r.get("hostname", ""),
            r.get("manufacturer", ""), r.get("http_port", ""),
            r.get("reachable", ""), r.get("login_ok", ""),
            r.get("rtsp_was_enabled", ""), r.get("api_was_enabled", ""),
            r.get("rtsp_enabled", ""), r.get("api_enabled", ""),
            r.get("rtsp_port", ""), r.get("channels", ""),
            r.get("action", ""), r.get("error", ""), status,
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col, value=v if v is not None else "")

        # Color the row by status
        fill = {
            "enabled": green, "already enabled": green,
            "skipped": gray, "unreachable": red,
            "login failed": red, "error": red,
        }.get(status, yellow)

        for col in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col).fill = fill

    # Auto-width
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

    wb.save(path)
    return True


def main():
    import argparse
    parser = argparse.ArgumentParser(description="Batch-enable RTSP + API on TVT NVRs")
    parser.add_argument("-p", "--password", default=NVR_PASSWORD, help="NVR admin password")
    parser.add_argument("-u", "--username", default="admin", help="NVR username (default: admin)")
    parser.add_argument("--timeout", type=int, default=10, help="Per-NVR timeout (default: 10s)")
    parser.add_argument("--dry-run", action="store_true", help="Only probe ports, don't enable")
    parser.add_argument("-o", "--output-dir", default=REPORT_DIR, help="Output directory for report")
    args = parser.parse_args()

    if not args.password:
        print("ERROR: Password required. Use -p or NVR_PASSWORD env var.", file=sys.stderr)
        sys.exit(1)

    # Load inventory
    with open(INVENTORY_PATH) as f:
        devices = json.load(f)

    # Deduplicate by IP (some sites list same NVR twice)
    seen_ips: dict[str, dict] = {}
    for d in devices:
        ip = d["ip"]
        if ip not in seen_ips:
            seen_ips[ip] = d

    report: list[dict] = []
    total = 0
    enabled_count = 0
    already_count = 0
    failed_count = 0
    skipped_count = 0
    unreachable_count = 0

    print(f"Loaded {len(devices)} entries ({len(seen_ips)} unique IPs)")
    print(f"{'='*90}")

    for ip, device in sorted(seen_ips.items()):
        site = device.get("site", "")
        hostname = device.get("hostname", "")
        mfr = device.get("manufacturer", "")

        entry = {
            "site": site, "ip": ip, "hostname": hostname,
            "manufacturer": mfr, "timestamp": datetime.now().isoformat(),
        }

        # Skip check
        skip_reason = should_skip(device)
        if skip_reason:
            entry.update({"status": "skipped", "action": f"skip: {skip_reason}",
                          "reachable": None, "login_ok": None})
            report.append(entry)
            skipped_count += 1
            print(f"  SKIP  {ip:<18} {site:<40} ({skip_reason})")
            continue

        total += 1
        print(f"  [{total:>3}]  {ip:<18} {site:<40} ", end="", flush=True)

        # Find HTTP port
        http_port = find_http_port(ip)
        if not http_port:
            entry.update({"status": "unreachable", "reachable": False,
                          "http_port": None, "error": "no HTTP port found"})
            report.append(entry)
            unreachable_count += 1
            print("UNREACHABLE")
            continue

        if http_port != 80:
            print(f"(port {http_port}) ", end="", flush=True)

        if args.dry_run:
            entry.update({"status": "dry-run", "reachable": True, "http_port": http_port})
            report.append(entry)
            print(f"port={http_port} (dry-run)")
            continue

        # Process
        result = process_nvr(ip, http_port, args.username, args.password, timeout=args.timeout)
        entry.update(result)

        if result.get("error"):
            if result.get("login_ok") is False:
                entry["status"] = "login failed"
                failed_count += 1
                print(f"LOGIN FAILED: {result['error']}")
            else:
                entry["status"] = "error"
                failed_count += 1
                print(f"ERROR: {result['error']}")
        elif result.get("action", "").startswith("already"):
            entry["status"] = "already enabled"
            already_count += 1
            ch = f" ({result['channels']} ch)" if result.get("channels") else ""
            print(f"ALREADY ON{ch}")
        else:
            entry["status"] = "enabled"
            enabled_count += 1
            ch = f" ({result['channels']} ch)" if result.get("channels") else ""
            print(f"ENABLED{ch} — {result.get('action', '')}")

        report.append(entry)

    # Summary
    print(f"\n{'='*90}")
    print(f"SUMMARY")
    print(f"  Total NVRs processed:  {total}")
    print(f"  Services enabled:      {enabled_count}")
    print(f"  Already enabled:       {already_count}")
    print(f"  Failed/errors:         {failed_count}")
    print(f"  Unreachable:           {unreachable_count}")
    print(f"  Skipped (non-TVT/110): {skipped_count}")

    # Save reports
    os.makedirs(args.output_dir, exist_ok=True)

    json_path = os.path.join(args.output_dir, "nvr_service_report.json")
    with open(json_path, "w") as f:
        json.dump(report, f, indent=2, default=str)
    print(f"\nJSON report: {json_path}")

    xlsx_path = os.path.join(args.output_dir, "nvr_service_report.xlsx")
    if save_xlsx(report, xlsx_path):
        print(f"XLSX report: {xlsx_path}")
    else:
        print("XLSX skipped (install openpyxl for Excel output)")


if __name__ == "__main__":
    main()
