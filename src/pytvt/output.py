"""Output formatters — console reports, CSV, JSON, and per-site XLSX.

All public functions accept :class:`~pytvt.models.ScanResult` and
:class:`~pytvt.models.DeviceEntry` objects — never raw dicts.
"""

from __future__ import annotations

import csv
import json
import sys
from collections import defaultdict
from pathlib import Path

from .models import DeviceEntry, ScanResult

# ── Console ──────────────────────────────────────────────────────────


def print_nvr_report(result: ScanResult) -> None:
    """Print a formatted console report for a single NVR scan."""
    header = f"\n{'=' * 80}"
    print(header)
    print(f"  Site: {result.site}")
    print(f"  NVR:  {result.hostname} ({result.nvr_ip})")
    print(f"{'=' * 80}")

    if not result.success:
        print(f"  ERROR: {result.error or 'Unknown error'}")
        return

    if result.device_name or result.device_model:
        print(f"  Device: {result.device_name} | Model: {result.device_model}")
    if result.firmware:
        print(f"  Firmware: {result.firmware}")
    if result.serial_number:
        print(f"  S/N: {result.serial_number}")
    print(f"  Total Channels: {result.total_channels}")
    print(f"  {'-' * 76}")

    cameras = [c for c in result.cameras if c.has_address]
    if cameras:
        print(f"  {'Ch':<4} {'Camera Name':<28} {'Address':<18} {'Port':<6} {'Status':<8} {'Model'}")
        print(f"  {'--':<4} {'-' * 27:<28} {'-' * 17:<18} {'----':<6} {'------':<8} {'-----'}")
        for cam in cameras:
            print(
                f"  {cam.channel!s:<4} {str(cam.name)[:27]:<28} "
                f"{cam.address:<18} {cam.port!s:<6} {cam.status:<8} {cam.model}"
            )
    else:
        print("  No camera details available")


def print_summary(results: list[ScanResult]) -> None:
    """Print a one-line scan summary."""
    total = len(results)
    ok = sum(1 for r in results if r.success)
    cams = sum(r.camera_count for r in results)
    print(f"\n{'=' * 80}")
    print(f"  SUMMARY: {ok}/{total} NVRs scanned successfully, {cams} cameras with addresses found")
    print(f"{'=' * 80}")


# ── CSV ──────────────────────────────────────────────────────────────


def save_csv(results: list[ScanResult], output_path: str) -> None:
    """Write all camera data to a flat CSV file."""
    rows = _flatten_results(results)
    if not rows:
        print("No data to save.", file=sys.stderr)
        return

    fieldnames = list(rows[0].keys())
    with open(output_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    print(f"\nCSV saved to: {output_path}")


# ── JSON ─────────────────────────────────────────────────────────────


def save_json(results: list[ScanResult], output_path: str) -> None:
    """Serialise all results to a JSON file."""
    with open(output_path, "w") as f:
        json.dump([r.to_dict() for r in results], f, indent=2)
    print(f"\nJSON saved to: {output_path}")


# ── XLSX ─────────────────────────────────────────────────────────────


def save_xlsx_per_site(results: list[ScanResult], output_dir: str) -> None:
    """Save scan results as one XLSX file per site with NVR Config and NVR Info tabs."""
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    sites: dict[str, list[ScanResult]] = defaultdict(list)
    for r in results:
        sites[r.site or "Unknown Site"].append(r)

    # Shared styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    style_kw = dict(
        header_font=header_font,
        header_fill=header_fill,
        header_align=header_align,
        thin_border=thin_border,
    )

    cam_columns = ["Camera Name", "Address", "Port", "Status", "Protocol", "Model"]
    nvr_columns = [
        "Hostname",
        "IP",
        "MAC",
        "Port",
        "Device Name",
        "Model",
        "Serial Number",
        "Firmware",
        "Channels",
        "Status",
    ]

    created: list[tuple[str, str, int, int]] = []

    for site_name, nvr_results in sorted(sites.items()):
        safe_name = "".join(c if c.isalnum() or c in " -_#()" else "_" for c in site_name).strip()
        xlsx_path = output_path / f"{safe_name}.xlsx"

        new_cam_count = sum(r.camera_count for r in nvr_results)

        # Protect existing data when the new scan got zero cameras
        if new_cam_count == 0 and xlsx_path.exists():
            existing_wb = load_workbook(xlsx_path)
            if "NVR Config" in existing_wb.sheetnames and existing_wb["NVR Config"].max_row > 1:
                old_count = existing_wb["NVR Config"].max_row - 1
                print(f"  {site_name}: scan returned 0 cameras, keeping existing file ({old_count} cameras)")
                existing_wb.close()
                continue
            existing_wb.close()

        wb = load_workbook(xlsx_path) if xlsx_path.exists() else Workbook()
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
            del wb["Sheet"]

        # ── NVR Info tab ─────────────────────────────────────────
        ws_info = _xlsx_upsert_sheet(wb, "NVR Info")
        _xlsx_write_header(ws_info, nvr_columns, **style_kw)

        row = 2
        for nvr in sorted(nvr_results, key=lambda r: r.hostname):
            status = "OK" if nvr.success else (nvr.error or "Failed")
            values = [
                nvr.hostname,
                nvr.nvr_ip,
                nvr.nvr_mac,
                nvr.nvr_port,
                nvr.device_name,
                nvr.device_model,
                nvr.serial_number,
                nvr.firmware,
                nvr.total_channels,
                status,
            ]
            for col_idx, value in enumerate(values, 1):
                cell = ws_info.cell(row=row, column=col_idx, value=value)
                cell.border = thin_border
            row += 1

        _xlsx_autofit(ws_info, nvr_columns, row - 1)
        if row > 2:
            ws_info.auto_filter.ref = f"A1:{get_column_letter(len(nvr_columns))}{row - 1}"

        # ── NVR Config (cameras) tab ─────────────────────────────
        ws_cam = _xlsx_upsert_sheet(wb, "NVR Config")
        _xlsx_write_header(ws_cam, cam_columns, **style_kw)

        cam_row = 2
        for nvr in nvr_results:
            for cam in nvr.cameras:
                if not cam.has_address:
                    continue
                values = [cam.name, cam.address, cam.port, cam.status, cam.protocol, cam.model]
                for col_idx, value in enumerate(values, 1):
                    cell = ws_cam.cell(row=cam_row, column=col_idx, value=value)
                    cell.border = thin_border
                cam_row += 1

        _xlsx_autofit(ws_cam, cam_columns, cam_row - 1)
        if cam_row > 2:
            ws_cam.auto_filter.ref = f"A1:{get_column_letter(len(cam_columns))}{cam_row - 1}"

        wb.save(xlsx_path)
        created.append((site_name, str(xlsx_path), row - 2, cam_row - 2))

    print(f"\nXLSX files saved to: {output_dir}")
    for site_name, fpath, nvrs, cams in created:
        print(f"  {site_name}: {nvrs} NVR(s), {cams} cameras -> {fpath}")


# ── Failed devices ───────────────────────────────────────────────────


def save_failed_devices(
    results: list[ScanResult],
    devices: list[DeviceEntry],
    output_path: str,
) -> None:
    """Write a JSON file of devices that failed to scan (for retry)."""
    failed_ips = {r.nvr_ip for r in results if not r.success}
    scanned_ips = {r.nvr_ip for r in results}
    missing_ips = {d.ip for d in devices} - scanned_ips

    all_failed_ips = failed_ips | missing_ips
    if not all_failed_ips:
        return

    failed = [d for d in devices if d.ip in all_failed_ips]
    with open(output_path, "w") as f:
        json.dump([_device_to_dict(d) for d in failed], f, indent=2)

    print(f"\n{len(failed)} failed/missing device(s) saved to: {output_path}")


# ── Private helpers ──────────────────────────────────────────────────


def _flatten_results(results: list[ScanResult]) -> list[dict]:
    """Flatten NVR + camera data into one row per camera for CSV export."""
    rows: list[dict] = []
    for nvr in results:
        nvr_fields = {
            "site": nvr.site,
            "nvr_hostname": nvr.hostname,
            "nvr_ip": nvr.nvr_ip,
            "nvr_mac": nvr.nvr_mac,
            "nvr_device_name": nvr.device_name,
            "nvr_model": nvr.device_model,
            "nvr_firmware": nvr.firmware,
            "nvr_serial": nvr.serial_number,
        }
        addressable = [c for c in nvr.cameras if c.has_address]
        if addressable:
            for cam in addressable:
                rows.append(
                    {
                        **nvr_fields,
                        "channel": cam.channel,
                        "camera_name": cam.name,
                        "camera_address": cam.address,
                        "camera_port": cam.port,
                        "camera_status": cam.status,
                        "camera_protocol": cam.protocol,
                        "camera_model": cam.model,
                    }
                )
        else:
            rows.append(
                {
                    **nvr_fields,
                    "channel": "",
                    "camera_name": "",
                    "camera_address": "",
                    "camera_port": "",
                    "camera_status": "",
                    "camera_protocol": "",
                    "camera_model": "",
                    "error": nvr.error or "",
                }
            )
    return rows


def _device_to_dict(d: DeviceEntry) -> dict:
    """Convert a DeviceEntry back to a JSON-serialisable dict."""
    out: dict = {"ip": d.ip}
    if d.site:
        out["site"] = d.site
    if d.hostname:
        out["hostname"] = d.hostname
    if d.mac:
        out["mac"] = d.mac
    if d.port:
        out["port"] = d.port
    if d.manufacturer:
        out["manufacturer"] = d.manufacturer
    return out


def _xlsx_upsert_sheet(wb, sheet_name: str):
    """Remove a sheet if it exists and create a fresh one."""
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    return wb.create_sheet(sheet_name)


def _xlsx_write_header(ws, columns, *, header_font, header_fill, header_align, thin_border):
    """Write a styled header row."""
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border


def _xlsx_autofit(ws, columns, last_row: int) -> None:
    """Auto-fit column widths based on content."""
    from openpyxl.utils import get_column_letter

    for col_idx in range(1, len(columns) + 1):
        max_len = len(columns[col_idx - 1])
        for row in range(2, last_row + 1):
            val = ws.cell(row=row, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 50)
