"""
TVT LAN Device Discovery — Pure Python SSDP/multicast implementation.

Discovers TVT NVRs, DVRs, IPCs and other devices on the local network by
sending an SSDP M-SEARCH multicast probe and parsing the XML responses.

Protocol details reverse-engineered from IPTool.app (macOS) and confirmed
against the official DVR_NET_SDK.h header (SEARCHED_DEVICE_INFO struct,
IPTool_SearchDataCallBack callback).

Discovery flow:
  1. Send M-SEARCH to 239.255.255.250:1900 with
     ST: urn:schemas-upnp-org:service:EmbeddedNetDeviceControl:1
  2. Each TVT device replies with an XML body (<multicastSearchResult>)
  3. Parse XML to extract device info (IP, MAC, model, ports, firmware, etc.)

Usage:
    from tvt_discovery import discover_devices
    devices = discover_devices(timeout=5)
    for dev in devices:
        print(dev["ip"], dev["product_model"], dev["device_name"])

CLI:
    python tvt_discovery.py [--timeout 5] [--json]
"""

import ipaddress
import json
import socket
import struct
import sys
import time
import xml.etree.ElementTree as ET
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Any

# ── SSDP constants ──────────────────────────────────────────────────────────

SSDP_ADDR = "239.255.255.250"
SSDP_PORT = 1900
SSDP_ST = "urn:schemas-upnp-org:service:EmbeddedNetDeviceControl:1"

SSDP_MSEARCH = (
    f'M-SEARCH * HTTP/1.1\r\nHOST:{SSDP_ADDR}:{SSDP_PORT}\r\nMan:"ssdp:discover"\r\nST:{SSDP_ST}\r\nMX:3\r\n\r\n'
)

# Device type mapping (from IPTool English language file + SDK header)
DEVICE_TYPES = {
    "TVT_DVR": "DVR",
    "TVT_NVR": "NVR",
    "TVT_IPC": "IPC",
    "TVT_MDVR": "MDVR",
    "TVT_STORAGE": "Storage",
    "TVT_DECODER": "Decoder",
    "TVT_NETKEYBOARD": "Network Keyboard",
}


def _get_local_ips() -> list[str]:
    """Get all local IPv4 addresses (non-loopback) for binding."""
    ips = []
    try:
        # Use UDP connect trick to find default route interface
        with socket.socket(socket.AF_INET, socket.SOCK_DGRAM) as s:
            s.connect(("8.8.8.8", 80))
            ips.append(s.getsockname()[0])
    except OSError:
        pass
    if not ips:
        ips.append("0.0.0.0")
    return ips


def _parse_xml_response(data: bytes, source_addr: tuple[str, int]) -> dict[str, Any] | None:
    """Parse a TVT multicast search XML response into a device dict.

    Expected XML structure:
        <multicastSearchResult>
          <tcpIp>
            <devName>NVR-01</devName>
            <ipAddr>192.168.1.100</ipAddr>
            <mask>255.255.255.0</mask>
            <maskAddr>255.255.255.0</maskAddr>
            <gateway>192.168.1.1</gateway>
            <dns1>8.8.8.8</dns1>
            <dns2>8.8.4.4</dns2>
            <macAddr>58:5B:69:XX:XX:XX</macAddr>
          </tcpIp>
          <port>
            <dataPort>9008</dataPort>
            <httpPort>80</httpPort>
          </port>
          <productInfo>
            <devName>NVR</devName>
            <productModel>TD-2708TS-C</productModel>
            <productSeries>N9000</productSeries>
            <softwareVer>V5.2.0</softwareVer>
            <kernelVer>...</kernelVer>
          </productInfo>
        </multicastSearchResult>
    """
    try:
        # The response may be raw XML or prefixed with HTTP headers.
        # Find the XML start.
        text = data.decode("utf-8", errors="replace")
        xml_start = text.find("<multicastSearchResult")
        if xml_start == -1:
            return None
        xml_text = text[xml_start:]

        root = ET.fromstring(xml_text)
    except ET.ParseError:
        return None

    def _get(parent_tag: str, child_tag: str) -> str:
        parent = root.find(parent_tag)
        if parent is None:
            return ""
        child = parent.find(child_tag)
        return (child.text or "").strip() if child is not None else ""

    def _get_root(tag: str) -> str:
        el = root.find(tag)
        return (el.text or "").strip() if el is not None else ""

    ip = _get("tcpIp", "ipAddr") or source_addr[0]
    mac = _get("tcpIp", "macAddr")
    device_name = _get("tcpIp", "devName") or _get("productInfo", "devName")
    product_model = _get("productInfo", "productModel")
    product_series = _get("productInfo", "productSeries")
    software_ver = _get("productInfo", "softwareVer")
    kernel_ver = _get("productInfo", "kernelVer")

    data_port_str = _get("port", "dataPort")
    http_port_str = _get("port", "httpPort")

    # Determine device type from product series or model
    device_type = "Unknown"
    series_upper = product_series.upper()
    for key, label in DEVICE_TYPES.items():
        if key.replace("TVT_", "") in series_upper:
            device_type = label
            break
    else:
        # Heuristic: check model prefix
        model_upper = product_model.upper()
        if model_upper.startswith("TD-") and ("NVR" in model_upper or "TS" in model_upper):
            device_type = "NVR"
        elif model_upper.startswith("TD-") and "DVR" in model_upper:
            device_type = "DVR"
        elif model_upper.startswith("IP-") or model_upper.startswith("TD-9"):
            device_type = "IPC"

    return {
        "ip": ip,
        "mac": mac,
        "device_name": device_name,
        "product_model": product_model,
        "product_series": product_series,
        "device_type": device_type,
        "software_version": software_ver,
        "kernel_version": kernel_ver,
        "data_port": int(data_port_str) if data_port_str.isdigit() else 0,
        "http_port": int(http_port_str) if http_port_str.isdigit() else 0,
        "subnet_mask": _get("tcpIp", "mask") or _get("tcpIp", "maskAddr"),
        "gateway": _get("tcpIp", "gateway"),
        "dns1": _get("tcpIp", "dns1"),
        "dns2": _get("tcpIp", "dns2"),
        "source_addr": source_addr[0],
    }


def discover_devices(
    timeout: float = 5.0,
    retries: int = 2,
    bind_addr: str | None = None,
) -> list[dict[str, Any]]:
    """Discover TVT devices on the LAN via SSDP multicast.

    Args:
        timeout: Seconds to wait for responses after each probe.
        retries: Number of M-SEARCH probes to send (devices may miss the first).
        bind_addr: Local IP to bind to (auto-detected if None).

    Returns:
        List of device dicts, de-duplicated by MAC address.
    """
    if bind_addr is None:
        local_ips = _get_local_ips()
        bind_addr = local_ips[0]

    seen: dict[str, dict] = {}  # keyed by MAC (or IP if no MAC)

    sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM, socket.IPPROTO_UDP)
    try:
        sock.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
        if hasattr(socket, "SO_REUSEPORT"):
            sock.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEPORT, 1)

        # Bind to any port on the chosen interface
        sock.bind((bind_addr, 0))

        # Set multicast TTL and outgoing interface
        sock.setsockopt(socket.IPPROTO_IP, socket.IP_MULTICAST_TTL, 4)
        sock.setsockopt(
            socket.IPPROTO_IP,
            socket.IP_MULTICAST_IF,
            socket.inet_aton(bind_addr),
        )

        msearch_bytes = SSDP_MSEARCH.encode("utf-8")

        for attempt in range(retries):
            # Send M-SEARCH probe
            sock.sendto(msearch_bytes, (SSDP_ADDR, SSDP_PORT))

            # Collect responses until timeout
            deadline = time.monotonic() + timeout
            while True:
                remaining = deadline - time.monotonic()
                if remaining <= 0:
                    break
                sock.settimeout(remaining)
                try:
                    data, addr = sock.recvfrom(65535)
                except socket.timeout:
                    break

                device = _parse_xml_response(data, addr)
                if device is None:
                    continue

                # De-duplicate by MAC, fall back to IP
                key = device["mac"] or device["ip"]
                if key not in seen:
                    seen[key] = device

    finally:
        sock.close()

    return sorted(seen.values(), key=lambda d: (d["device_type"], d["ip"]))


# ── TVT protocol constants for TCP probe ────────────────────────────────────

_TVT_HEADER_FLAG = b"1111"
_TVT_DATA_PORT = 9008
_TVT_INIT_FLAG = b"head"


def _tcp_probe_tvt(ip: str, port: int = _TVT_DATA_PORT, timeout: float = 2.0) -> dict[str, Any] | None:
    """Try a TCP connection to a TVT data port and check for the TVT handshake.

    Returns a minimal device dict if the host responds with a TVT init packet,
    or None if it's not a TVT device / unreachable.
    """
    try:
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            s.settimeout(timeout)
            s.connect((ip, port))
            # TVT devices send an init packet immediately after TCP connect
            data = s.recv(256)
            if _TVT_INIT_FLAG in data or _TVT_HEADER_FLAG in data:
                return {
                    "ip": ip,
                    "mac": "",
                    "device_name": "",
                    "product_model": "",
                    "product_series": "",
                    "device_type": "Unknown",
                    "software_version": "",
                    "kernel_version": "",
                    "data_port": port,
                    "http_port": 0,
                    "subnet_mask": "",
                    "gateway": "",
                    "dns1": "",
                    "dns2": "",
                    "source_addr": ip,
                    "_probe": "tcp",
                }
    except (OSError, socket.timeout):
        pass
    return None


def _unicast_probe(ip: str, timeout: float = 2.0) -> dict[str, Any] | None:
    """Send a unicast M-SEARCH to a single IP on port 1900 and parse the response."""
    msearch = SSDP_MSEARCH.encode("utf-8")
    try:
        with socket.socket(socket.AF_INET, socket.SOCK_DGRAM, socket.IPPROTO_UDP) as s:
            s.settimeout(timeout)
            s.sendto(msearch, (ip, SSDP_PORT))
            data, addr = s.recvfrom(65535)
            device = _parse_xml_response(data, addr)
            if device:
                device["_probe"] = "unicast"
            return device
    except (OSError, socket.timeout):
        return None


def discover_subnet(
    cidr: str,
    timeout: float = 2.0,
    concurrency: int = 50,
    tcp_fallback: bool = True,
    tcp_port: int = _TVT_DATA_PORT,
    progress: bool = True,
) -> list[dict[str, Any]]:
    """Discover TVT devices on a remote (routed) subnet via unicast probes.

    Sends a unicast UDP M-SEARCH to each host in the CIDR range.  Hosts that
    don't reply to UDP get a TCP port probe on the TVT data port as fallback.

    Args:
        cidr: Subnet in CIDR notation, e.g. "10.200.50.0/24".
        timeout: Per-host probe timeout in seconds.
        concurrency: Max parallel probes.
        tcp_fallback: Also try TCP port probe for non-responders.
        tcp_port: TVT data port for TCP probe (default 9008).
        progress: Print progress to stderr.

    Returns:
        List of discovered device dicts.
    """
    network = ipaddress.ip_network(cidr, strict=False)
    hosts = [str(h) for h in network.hosts()]  # excludes network & broadcast

    if progress:
        print(f"Sweeping {len(hosts)} hosts in {network} (concurrency={concurrency})...", file=sys.stderr)

    seen: dict[str, dict] = {}
    udp_misses: list[str] = []

    # Phase 1: unicast UDP M-SEARCH
    with ThreadPoolExecutor(max_workers=concurrency) as pool:
        futures = {pool.submit(_unicast_probe, ip, timeout): ip for ip in hosts}
        done_count = 0
        for future in as_completed(futures):
            done_count += 1
            ip = futures[future]
            try:
                device = future.result()
                if device:
                    key = device["mac"] or device["ip"]
                    if key not in seen:
                        seen[key] = device
                        if progress:
                            print(
                                f"  [{done_count}/{len(hosts)}] {ip} — FOUND ({device['product_model'] or 'TVT device'})",
                                file=sys.stderr,
                            )
                else:
                    udp_misses.append(ip)
            except Exception:
                udp_misses.append(ip)

    # Phase 2: TCP port probe fallback
    if tcp_fallback and udp_misses:
        if progress:
            print(f"UDP: {len(seen)} found. TCP fallback on {len(udp_misses)} remaining hosts...", file=sys.stderr)
        with ThreadPoolExecutor(max_workers=concurrency) as pool:
            futures = {pool.submit(_tcp_probe_tvt, ip, tcp_port, timeout): ip for ip in udp_misses}
            for future in as_completed(futures):
                ip = futures[future]
                try:
                    device = future.result()
                    if device:
                        key = device["ip"]
                        if key not in seen:
                            seen[key] = device
                            if progress:
                                print(f"  {ip} — FOUND via TCP (port {tcp_port})", file=sys.stderr)
                except Exception:
                    pass

    return sorted(seen.values(), key=lambda d: (d["device_type"], d["ip"]))


def print_discovery_report(devices: list[dict]) -> None:
    """Print a formatted table of discovered devices."""
    if not devices:
        print("No TVT devices found on the network.")
        return

    print(f"\n{'=' * 100}")
    print(f"  TVT LAN Device Discovery — {len(devices)} device(s) found")
    print(f"{'=' * 100}")
    print(f"  {'Type':<8} {'IP Address':<18} {'MAC':<20} {'Model':<20} {'Name':<20} {'FW':<12} {'Ports'}")
    print(f"  {'----':<8} {'-' * 17:<18} {'-' * 19:<20} {'-' * 19:<20} {'-' * 19:<20} {'-' * 11:<12} {'-----'}")
    for d in devices:
        ports = []
        if d["data_port"]:
            ports.append(f"data:{d['data_port']}")
        if d["http_port"]:
            ports.append(f"http:{d['http_port']}")
        print(
            f"  {d['device_type']:<8} {d['ip']:<18} {d['mac']:<20} "
            f"{d['product_model']:<20} {d['device_name'][:19]:<20} "
            f"{d['software_version'][:11]:<12} {', '.join(ports)}"
        )
    print()


def save_discovery_xlsx(devices: list[dict], output_path: str) -> None:
    """Save discovery results to an XLSX file with a formatted 'Discovered Devices' sheet."""
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from pathlib import Path

    columns = [
        ("Type", 10),
        ("IP Address", 18),
        ("MAC", 20),
        ("Model", 22),
        ("Name", 22),
        ("Firmware", 20),
        ("Data Port", 12),
        ("HTTP Port", 12),
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    if path.exists():
        wb = load_workbook(path)
    else:
        wb = Workbook()

    sheet_name = "Discovered Devices"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name, 0)

    # Remove default empty sheet if it exists
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    # Header
    for col_idx, (name, _) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows
    for row_idx, d in enumerate(devices, 2):
        values = [
            d["device_type"],
            d["ip"],
            d["mac"],
            d["product_model"],
            d["device_name"],
            d["software_version"],
            d["data_port"] or "",
            d["http_port"] or "",
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border

    # Auto-fit column widths
    for col_idx, (_, width) in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(devices) + 1}"
    ws.freeze_panes = "A2"

    wb.save(path)
    print(f"Discovery results ({len(devices)} devices) saved to: {path}")


def discovery_to_scanner_format(devices: list[dict], site: str = "Discovered", scan_port: int = 6036) -> list[dict]:
    """Convert discovery results to the JSON format expected by pytvt scanner.

    This allows piping discovered devices directly into main.py for a full scan.
    All discovered devices are included (NVR, DVR, IPC, Unknown).

    Args:
        devices: List of discovered device dicts.
        site: Site name to assign to each device.
        scan_port: Protocol port to assign (default 6036 — the TVT Server Port).
    """
    scanner_devices = []
    for d in devices:
        scanner_devices.append(
            {
                "ip": d["ip"],
                "port": scan_port,
                "data_port": d["data_port"] or 9008,
                "http_port": d["http_port"] or 0,
                "mac": d["mac"],
                "hostname": d["device_name"] or d["product_model"] or d["ip"],
                "site": site,
                "manufacturer": "TVT",
                "model": d["product_model"],
                "firmware": d["software_version"],
                "device_type": d["device_type"],
                "serial": "",
                "_discovered": True,
            }
        )
    return scanner_devices


def main():
    import argparse

    parser = argparse.ArgumentParser(
        description="TVT Device Discovery — find TVT NVRs/IPCs via multicast or subnet sweep"
    )
    parser.add_argument(
        "--timeout",
        "-t",
        type=float,
        default=5.0,
        help="Seconds to wait for responses per probe (default: 5)",
    )
    parser.add_argument(
        "--retries",
        "-r",
        type=int,
        default=2,
        help="Number of M-SEARCH probes to send in multicast mode (default: 2)",
    )
    parser.add_argument(
        "--bind",
        help="Local IP address to bind to (auto-detected if omitted)",
    )
    parser.add_argument(
        "--subnet",
        action="append",
        metavar="CIDR",
        help="Scan a remote subnet via unicast probes (e.g. 10.200.50.0/24). Can be specified multiple times.",
    )
    parser.add_argument(
        "--concurrency",
        type=int,
        default=50,
        help="Max parallel probes for subnet sweep (default: 50)",
    )
    parser.add_argument(
        "--no-tcp-fallback",
        action="store_true",
        help="Skip TCP port-probe fallback during subnet sweep",
    )
    parser.add_argument(
        "--tcp-port",
        type=int,
        default=_TVT_DATA_PORT,
        help=f"TVT data port for TCP probe fallback (default: {_TVT_DATA_PORT})",
    )
    parser.add_argument(
        "--json",
        action="store_true",
        help="Output raw JSON instead of formatted table",
    )
    parser.add_argument(
        "--scanner-json",
        metavar="FILE",
        help="Write discovered NVRs in pytvt scanner JSON format to FILE",
    )
    parser.add_argument(
        "--site",
        default="Discovered",
        help="Site name to assign when generating scanner JSON (default: Discovered)",
    )
    parser.add_argument(
        "--xlsx",
        metavar="FILE",
        help="Save discovery results to an XLSX file",
    )
    parser.add_argument(
        "--scan-port",
        type=int,
        default=6036,
        help="Protocol port to use in scanner JSON output (default: 6036)",
    )
    args = parser.parse_args()

    all_devices: list[dict] = []

    # Subnet sweep mode
    if args.subnet:
        for cidr in args.subnet:
            devices = discover_subnet(
                cidr,
                timeout=args.timeout,
                concurrency=args.concurrency,
                tcp_fallback=not args.no_tcp_fallback,
                tcp_port=args.tcp_port,
            )
            all_devices.extend(devices)
    else:
        # Default: LAN multicast
        print(f"Searching for TVT devices on LAN (timeout={args.timeout}s, retries={args.retries})...")
        all_devices = discover_devices(
            timeout=args.timeout,
            retries=args.retries,
            bind_addr=args.bind,
        )

    devices = all_devices

    if args.json:
        print(json.dumps(devices, indent=2))
    else:
        print_discovery_report(devices)

    if args.scanner_json:
        scanner_devices = discovery_to_scanner_format(devices, site=args.site, scan_port=args.scan_port)
        with open(args.scanner_json, "w") as f:
            json.dump(scanner_devices, f, indent=2)
        print(f"Scanner JSON ({len(scanner_devices)} devices) saved to: {args.scanner_json}")

    if args.xlsx:
        save_discovery_xlsx(devices, args.xlsx)

    return devices


if __name__ == "__main__":
    main()
